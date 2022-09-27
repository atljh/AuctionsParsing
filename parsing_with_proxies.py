from datetime import datetime, timedelta
from typing import Dict, List
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font 
import config
from openpyxl import load_workbook


headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64; rv:104.0) Gecko/20100101 Firefox/104.0',
}

proxies = {
    'https': f'http://{config.proxy_login}:{config.proxy_password}@{config.proxy_host}:{config.proxy_port}', 
}


def init_excel():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(('Auction Status', 'Auction Date', 'Amount', 'Sold To','Auction Type', 'Case #', 'Final Judgement Amount', 'Parcel ID', 'Property Address', 'Assessed Value', 'Plaintiff Max Bid'))
    for row in 'ABCDEFGHIJKL':
        if row == 'I':
            sheet.column_dimensions[row].width = 40
        else:
            sheet.column_dimensions[row].width = 25
    for row in range(1, 12):
        sheet.cell(row = 1, column = row).font = Font(size=15)
    workbook.save(filename=f"{config.filename}")


def excel(auctions: List[Dict[str, str]]) -> None:
    workbook = load_workbook(filename=f'{config.filename}')
    sheet = workbook.active
    for auc in auctions:
        sheet.append(tuple(auc.values()))
    workbook.save(filename=f"{config.filename}")


def get_location(url: str) -> None:
    response  = requests.get(url=url, headers=headers, proxies=proxies)   
    soup = BeautifulSoup(response.text, 'lxml')

    ip = soup.find('div', class_='ip').text.strip()
    location = soup.find('div', class_='value-country').text.strip()
    print(f'IP: {ip}\nLocation: {location}')


def parse(date: str) -> None:
    url = f'https://lee.realforeclose.com/index.cfm?zaction=AUCTION&Zmethod=PREVIEW&AUCTIONDATE={date}'
    response = requests.get(url=url, headers=headers, proxies=proxies)
    auctions_list = []

    if response.status_code == '403':
        print("Wrong proxy")
        return response.status_code
    
    soup = BeautifulSoup(response.text, 'lxml')
    auctions = soup.find_all('div', class_="AUCTION_ITEM")
    for auc in auctions:
        auction_status = auc.find('div', class_="ASTAT_MSGA").text 
        auction_date = auc.find('div', class_="ASTAT_MSGB").text
        amount = auc.find('div', class_="ASTAT_MSGD").text
        sold_to = auc.find('div', class_="ASTAT_MSG_SOLDTO_MSG").text

        if auction_status == 'Auction Status':
            context = {'Auction Status': auction_date, 'Auction Date': '', 'Amount': amount, 'Sold To': sold_to}
        else:
            context = {'Auction Status': auction_status, 'Auction Date': auction_date, 'Amount': amount, 'Sold To': sold_to}

        for table in auc.find('table', class_="ad_tab"):
            for label, value in zip(table.find_all('th', class_="AD_LBL"), table.find_all('td', class_="AD_DTA")):
                lbl = label.text.replace(':', '')
                if lbl == '':
                    context['Property Address'] += value.text
                else:
                    context[lbl] = value.text
        auctions_list.append(context)
    excel(auctions_list)


def date_generator(date_from: str, date_to: str) -> None:
    date_to = datetime.strptime(date_to, '%m/%d/%Y')
    current_date = datetime.strptime(date_from, '%m/%d/%Y')
    while current_date != date_to:
        current_date += timedelta(days=1)
        parse(current_date.strftime('%m/%d/%Y'))


def main():
    init_excel()
    get_location('https://2ip.io/')
    date_generator(date_from=config.date_from, date_to=config.date_to)




if __name__ == '__main__':
    main()

